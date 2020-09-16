import xlrd
from datetime import date,datetime
from xlrd import xldate_as_tuple
import xlsxwriter
import math
import copy
from openpyxl import load_workbook

#读取原始数据
excel_name = r'.\附件1：123家有信贷记录企业的相关数据.xlsx'
#原始数据中有一些是文本格式，Python读不了，需要预先在Excel中处理一下
book = xlrd.open_workbook(excel_name)
sheet_out = book.sheet_by_name('销项发票信息')
sheet_in = book.sheet_by_name('进项发票信息')

#用于保存各企业的在Excel中的位置信息
#OUT为销项信息，IN为进项信息
location_in = {}
location_out = {}           
copr_count = 0              

#表格总的行数
nrows_out = sheet_out.nrows
nrows_in = sheet_in.nrows

#求出各企业的在Excel中的位置信息：销项
for line in range(1,nrows_out):
    if sheet_out.cell(line,0).value != sheet_out.cell(line-1,0).value:
        copr_count += 1
        location_out['E'+str(copr_count)] = [line,0]
        if line != 1:
            location_out['E'+str(copr_count-1)][1] = line-1
    if line == nrows_out-1:
        location_out['E'+str(copr_count)][1] = nrows_out-1

copr_count = 0

#求出各企业的在Excel中的位置信息：进项
for line in range(1,nrows_in):
    if sheet_in.cell(line,0).value != sheet_in.cell(line-1,0).value:
        copr_count += 1
        location_in['E'+str(copr_count)] = [line,0]
        if line != 1:
            location_in['E'+str(copr_count-1)][1] = line-1
    if line == nrows_in-1:
        location_in['E'+str(copr_count)][1] = nrows_in-1

#一个用于求平均值的函数
def mean(data):
    b = len(data)
    sum = 0
    for i in data:
        sum = sum +i
    return sum/b

#用于将输出的公司信息保存在Excel中
def write_down_abs(corp_name,data_pack):
    excel_name = r'.\123家公司信息统计.xlsx'    #这个需要预先创建好，同时程序没有添加表头
    record_book = load_workbook(excel_name)
    sheetnames = record_book.sheetnames
    sheet = record_book[sheetnames[0]]

    row = str(int(corp_name[1:])+1)
#data_pack各项含义：
    sheet['A'+row] = corp_name
    sheet['B'+row] = data_pack[0]
    sheet['C'+row] = data_pack[1]
    sheet['D'+row] = data_pack[2]
    sheet['E'+row] = data_pack[3]
    sheet['F'+row] = data_pack[4]
    sheet['G'+row] = data_pack[5]
    sheet['H'+row] = data_pack[6]
    sheet['I'+row] = data_pack[7]
    sheet['J'+row] = data_pack[8]

    print('Writng: '+corp_name)
    record_book.save(excel_name)
    return 0

def one_copr_handler(copr_name,rows_out,rows_in):
    tuikuan_out = {}            #销项退款次数
    zuofeifapiao_out = {}       #销项作废发票数
    shuie_out = {}              #销项税额
    jine_out = {}               #销项金额
    jiaoyishu_out = {}          #销项交易数目

    tuikuan_in = {}             #进项退款次数
    zuofeifapiao_in = {}        #进项作废发票数
    shuie_in = {}               #进项税额
    jine_in = {}                #进项金额
    jiaoyishu_in = {}           #进项交易数目

    #销项发票的日期数据处理
    #起始月份的处理：若一家公司的信息记录起始于2月，我们把它归到第一节度，1月
    first_month_raw_out = int(str(datetime(*xldate_as_tuple(sheet_out.cell(rows_out[0], 2).value, 0)))[5:7])
    if first_month_raw_out >= 1 and first_month_raw_out <= 3:
        first_month_out = 1
    if first_month_raw_out >= 4 and first_month_raw_out <= 6:
        first_month_out = 4
    if first_month_raw_out >= 7 and first_month_raw_out <= 9:
        first_month_out = 7
    if first_month_raw_out >= 10 and first_month_raw_out <= 12:
        first_month_out = 10

    last_month_raw_out = int(str(datetime(*xldate_as_tuple(sheet_out.cell(rows_out[1], 2).value, 0)))[5:7])
    if last_month_raw_out >= 1 and last_month_raw_out <= 3:
        last_month_out = 3
    if last_month_raw_out >= 4 and last_month_raw_out <= 6:
        last_month_out = 6
    if last_month_raw_out >= 7 and last_month_raw_out <= 9:
        last_month_out = 9
    if last_month_raw_out >= 10 and last_month_raw_out <= 12:
        last_month_out = 12

    #进项发票的日期数据处理
    first_month_raw_in = int(str(datetime(*xldate_as_tuple(sheet_in.cell(rows_in[0], 2).value, 0)))[5:7])                  #获得最开始的月份
    if first_month_raw_in >= 1 and first_month_raw_in <= 3:
        first_month_in = 1
    if first_month_raw_in >= 4 and first_month_raw_in <= 6:
        first_month_in = 4
    if first_month_raw_in >= 7 and first_month_raw_in <= 9:
        first_month_in = 7
    if first_month_raw_in >= 10 and first_month_raw_in <= 12:
        first_month_in = 10

    last_month_raw_in = int(str(datetime(*xldate_as_tuple(sheet_in.cell(rows_in[1], 2).value, 0)))[5:7])
    if last_month_raw_in >= 1 and last_month_raw_in <= 3:
        last_month_in = 3
    if last_month_raw_in >= 4 and last_month_raw_in <= 6:
        last_month_in = 6
    if last_month_raw_in >= 7 and last_month_raw_in <= 9:
        last_month_in = 9
    if last_month_raw_in >= 10 and last_month_raw_in <= 12:
        last_month_in = 12

    first_year_out = int(str(datetime(*xldate_as_tuple(sheet_out.cell(rows_out[0],2).value, 0)))[0:4])
    first_year_in = int(str(datetime(*xldate_as_tuple(sheet_in.cell(rows_in[0],2).value, 0)))[0:4])

    #对销项发票信息的处理及收集
    for i in range(rows_out[0],rows_out[1]+1):
        delta_year = int(str(datetime(*xldate_as_tuple(sheet_out.cell(i,2).value, 0)))[0:4]) - first_year_out
        month = int(str(datetime(*xldate_as_tuple(sheet_out.cell(i,2).value, 0)))[5:7])
        #计算从起始时间开始算经过的季度
        sesson_count = ((delta_year*12 + month - first_month_out) // 3) + 1
        
        #这一季度信息还未被创建时
        if not 'S'+str(sesson_count) in tuikuan_out:
            tuikuan_out['S'+str(sesson_count)] = 0
            zuofeifapiao_out['S'+str(sesson_count)] = 0
            shuie_out['S'+str(sesson_count)] = 0
            jine_out['S'+str(sesson_count)] = 0
            jiaoyishu_out['S'+str(sesson_count)] = 0
        
        #计算之前所述的5项指标：
        if sheet_out.cell(i, 4).value < 0:
            if 'S'+str(sesson_count) in tuikuan_out:
                tuikuan_out['S'+str(sesson_count)] += 1

        if sheet_out.cell(i, 7).value == '作废发票':
            if 'S'+str(sesson_count) in zuofeifapiao_out:
                zuofeifapiao_out['S'+str(sesson_count)] += 1

        shuie_out['S'+str(sesson_count)] += sheet_out.cell(i, 5).value

        jine_out['S'+str(sesson_count)] += sheet_out.cell(i, 4).value

        jiaoyishu_out['S'+str(sesson_count)] += 1

        #数据结尾的处理
        if i == rows_out[1] + 1 - 1:
            multi_end = 3/(last_month_raw_out-last_month_out+3)
            tuikuan_out['S'+str(sesson_count)] *= multi_end
            zuofeifapiao_out['S'+str(sesson_count)] *= multi_end
            shuie_out['S'+str(sesson_count)] *= multi_end
            jine_out['S'+str(sesson_count)] *= multi_end
            jiaoyishu_out['S'+str(sesson_count)] *= multi_end

        #开头的处理
            multi_head = 3/(first_month_out+2-first_month_raw_out +1)
            tuikuan_out['S1'] *= multi_head
            zuofeifapiao_out['S1'] *= multi_head
            shuie_out['S1'] *= multi_head
            jine_out['S1'] *= multi_head
            jiaoyishu_out['S1'] *= multi_head

            sesson_out = sesson_count



    #对进项发票信息的处理和收集
    for i in range(rows_in[0],rows_in[1]+1):
        delta_year = int(str(datetime(*xldate_as_tuple(sheet_in.cell(i,2).value, 0)))[0:4]) - first_year_in
        month = int(str(datetime(*xldate_as_tuple(sheet_in.cell(i,2).value, 0)))[5:7])
        #计算从起始时间开始算经过的季度
        sesson_count = ((delta_year*12 + month - first_month_in) // 3) + 1

        #这一季度信息还未被创建时
        if not 'S'+str(sesson_count) in tuikuan_in:
            tuikuan_in['S'+str(sesson_count)] = 0
            zuofeifapiao_in['S'+str(sesson_count)] = 0
            shuie_in['S'+str(sesson_count)] = 0
            jine_in['S'+str(sesson_count)] = 0
            jiaoyishu_in['S'+str(sesson_count)] = 0
        
        #计算之前所述的5项指标：
        if sheet_in.cell(i, 4).value < 0:
            if 'S'+str(sesson_count) in tuikuan_in:
                tuikuan_in['S'+str(sesson_count)] += 1

        if sheet_in.cell(i, 7).value == '作废发票':
            if 'S'+str(sesson_count) in zuofeifapiao_in:
                zuofeifapiao_in['S'+str(sesson_count)] += 1

        shuie_in['S'+str(sesson_count)] += sheet_in.cell(i, 5).value

        jine_in['S'+str(sesson_count)] += sheet_in.cell(i, 4).value

        jiaoyishu_in['S'+str(sesson_count)] += 1

        #数据结尾的处理
        if i == rows_in[1] + 1 - 1:
            multi_end = 3/(last_month_raw_in-last_month_in+3)
            tuikuan_in['S'+str(sesson_count)] *= multi_end
            zuofeifapiao_in['S'+str(sesson_count)] *= multi_end
            shuie_in['S'+str(sesson_count)] *= multi_end
            jine_in['S'+str(sesson_count)] *= multi_end
            jiaoyishu_in['S'+str(sesson_count)] *= multi_end
        #开头的处理
            multi_head = 3/(first_month_in+2-first_month_raw_in +1)
            tuikuan_in['S1'] *= multi_head
            zuofeifapiao_in['S1'] *= multi_head
            shuie_in['S1'] *= multi_head
            jine_in['S1'] *= multi_head
            jiaoyishu_in['S1'] *= multi_head

            sesson_in = sesson_count

    '''
    print(tuikuan_out)
    print(zuofeifapiao_out)
    print(shuie_out)
    print(jine_out)
    print(jiaoyishu_out)

    print(tuikuan_in)
    print(zuofeifapiao_in)
    print(shuie_in)
    print(jine_in)
    print(jiaoyishu_in)
    '''
    '''
    print(copr_name,'进项季度数:',sesson_in)
    print(copr_name,'销项季度数:',sesson_out)
    
    print(copr_name+':',jine_out)
    print(copr_name+':',shuie_in)
    '''
    #销售额增长率
    sale_gain = copy.deepcopy(jine_out)
    for i in range(1,sesson_out+1):
        key = 'S'+str(i)
        if i == 1:
            sale_gain[key] = 0
        else:
            try:
                sale_gain[key] = (jine_out[key]-jine_out['S'+str(i-1)])/jine_out[key]
            except: #ZeroDivisionError:
                sale_gain[key] = 0
    
    #利润率
    if sesson_in >= sesson_out:
        profit_rate = copy.deepcopy(jine_out)
    else:
        profit_rate = copy.deepcopy(jine_in)
    for key in profit_rate:
        try:
            profit_rate[key] = (jine_out[key]-jine_in[key])/jine_out[key]
        except:# ZeroDivisionError:
            profit_rate[key] = 0

    #利润增长率
    profit_rate_gain = copy.deepcopy(profit_rate)
    for i in range(1,len(profit_rate_gain.keys())+1):
        key = 'S'+str(i)
        if i == 1:
            profit_rate_gain[key] = 0
        else:
            try:
                profit_rate_gain[key] = (profit_rate[key]-profit_rate['S'+str(i-1)])/profit_rate[key]
            except: #ZeroDivisionError:
                profit_rate_gain[key] = 0

    #退货率
    return_rate_out = copy.deepcopy(jiaoyishu_out)
    for key in return_rate_out:
        return_rate_out[key] = (tuikuan_out[key]+zuofeifapiao_out[key])/jiaoyishu_out[key]

    return_rate_in = copy.deepcopy(jiaoyishu_in)
    for key in return_rate_in:
        return_rate_in[key] = (tuikuan_in[key]+zuofeifapiao_in[key])/jiaoyishu_in[key]

    #最大金额流水
    max_jine_out = max(jine_out.values())

    #税额
    if sesson_in >= sesson_out:
        tax = copy.deepcopy(shuie_out)
        for key in tax:
            try:
                tax[key] = shuie_out[key]-shuie_in[key]
            except:
                tax[key] = 0
    else:
        tax = copy.deepcopy(shuie_in)
        for key in tax:
            try:
                tax[key] = shuie_out[key]-shuie_in[key]
            except:
                tax[key] = 0

        

    data_pack = [mean(sale_gain.values())/100,mean(profit_rate.values())/100,mean(profit_rate_gain.values())/100 \
    ,mean(return_rate_out.values())/100,mean(return_rate_in.values())/100,max_jine_out,max(mean(tax.values()),0) \
    ,mean(jine_out.values()),mean(jine_in.values())]

    #保存数据到Excel
    write_down_abs(copr_name,data_pack)
    
    return 0

if __name__ == "__main__":
    for key in location_in:
        one_copr_handler(key,location_out[key],location_in[key])
        print('Handlering: '+key)